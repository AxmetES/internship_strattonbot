import names
from openpyxl import Workbook
from datetime import date
from datetime import datetime

from utils import make_skcu_folder


def generate_xlsx_file():
    current_date = date.today()
    wb = Workbook()

    for i in range(0, 3):
        ws = wb.create_sheet(f"TDSheet{i}")
        ws['A1'] = 'Name'
        ws['B1'] = 'Current date'
        ws['C1'] = 'Current time'
        for j in range(2, 100):
            ws['A' + str(j)] = names.get_full_name()
            ws['B' + str(j)] = current_date
            ws['C' + str(j)] = datetime.now().strftime('%H:%M:%S')

    skcu_folder = make_skcu_folder()
    file_name = f'''strattonbot_{current_date}_{datetime.now().strftime('%H:%M:%S')}.xlsx'''
    file_dir = str(skcu_folder / file_name)
    wb.save(file_dir)
    return file_dir

if __name__ == "__main__":
    generate_xlsx_file()